import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import re
import time


def read_words(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    words = []
    for row in sheet.iter_rows(min_row=1, max_col=1):
        cell_value = row[0].value
        if cell_value is None:
            break
        words.append(cell_value)
    return words


def search_google(driver, query):
    search_url = f"https://www.google.com/search?q={query}"
    driver.get(search_url)
    time.sleep(0.1)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    return soup


def find_digikala_link(soup):
    if soup is None:
        return '-'
    try:
        links = soup.find_all('a', href=True)
        for link in links:
            href = link['href']
            if 'digikala.com' in href:
                return href
    except Exception as e:
        print(f"Error while finding link: {e}")
    return '-'


def save_results(results, output_file):
    wb = openpyxl.Workbook()
    sheet = wb.active

    for index, result in enumerate(results, start=1):
        sheet.cell(row=index, column=1).value = result

    try:
        wb.save(output_file)
        print(f"Results successfully saved to {output_file}")
    except Exception as e:
        print(f"Error saving results to {output_file}: {e}")


def contains_model_and_number(word):
    pattern = re.compile(r'مدل\s*\d+')
    return bool(pattern.search(word))


def contains_model_any_number(word):
    pattern = re.compile(r'مدل.*\d+')
    return bool(pattern.search(word))


def main(input_file, output_file):
    start_time = time.time()

    words = read_words(input_file)
    results = []
    foundLink = 0
    totalLink = 0

    options = Options()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    options.add_argument(f'user-agent={user_agent}')
    driver = webdriver.Chrome(service=Service(
        ChromeDriverManager().install()), options=options)

    try:
        for word in words:
            totalLink += 1
            if any(keyword in word for keyword in ["کد", "ساختنی مدل", "گوشی", "ظرفیت", "گیگابایت", "سایز", "سانتی متر", "سانتیمتر", "ارتفاع", "چند رنگ", "عددی", "سامسونگ", "اپل"]) or contains_model_and_number(word) or contains_model_any_number(word):
                foundLink += 1
                print(f"Searching for: {word}")
                soup = search_google(driver, word)
                digikala_link = find_digikala_link(soup)
                results.append(digikala_link)

                if digikala_link == '-':
                    print(f"Link not found for: {word}")
            elif any(keyword in word for keyword in ["انگشتر", "عقیق"]):
                results.append("یافت نشد")
                print(f"No result for: {word}")
            else:
                results.append("")
                print(f"Skipped: {word}")

            time.sleep(0.1)

    except Exception as e:
        print(f"Error during the main process: {e}")
    finally:
        driver.quit()

    save_results(results, output_file)

    end_time = time.time()
    elapsed_time = end_time - start_time

    print(f"Results saved to {output_file}")
    print(f"{foundLink}/{totalLink} link(s) found")
    # Print total time taken
    print(f"Total time taken: {elapsed_time:.2f} seconds")


if __name__ == "__main__":
    input_file = 'words.xlsx'
    output_file = 'results.xlsx'
    main(input_file, output_file)
